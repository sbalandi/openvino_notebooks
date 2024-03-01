import argparse

import openpyxl
import csv
from github import Github  # type: ignore
from pathlib import Path

from datetime import datetime
import requests
import base64

import logging

logging.basicConfig()
logger = logging.getLogger('statistic')
logger.setLevel(logging.INFO)


URL_PATH_KEY = 'url_path'
REFERRER_KEY = 'referrer'
VIEWS_UNIQUE_KEY = 'views_unique'
VIEWS_TOTAL_KEY = 'views_total'

class Contents():
    def __init__(self, decoded_content, name):
        self.decoded_content = decoded_content
        self.name = name

def get_contents_from_local_dir(dir_path):
    dir_path = Path(dir_path)
    contents = []

    for i in dir_path.glob('**/*'):
        if i.is_file():
            contents.append(Contents(i.read_bytes(), i.name))

    return contents


def create_exel(paths_data, referrs_data, report_name: Path):
    logger.info(f'Create xslx file')
    exel_builed = ExelBuiler()

# ------------------------------------- pivot_paths ------------------------------------
    exel_builed.create_content_sheet('paths_content', paths_data)

# ------------------------------------- paths ------------------------------------
    exel_builed.create_paths_snapshoot_sheet('paths_spanshots', paths_data)

# ------------------------------------- pivot_sources ------------------------------------
    exel_builed.create_content_sheet('referrers_content', referrs_data)

# ------------------------------------- sources ------------------------------------
    exel_builed.create_paths_snapshoot_sheet('referrers_spanshots', referrs_data)

# ------------------------------------- paths_chart ------------------------------------
    exel_builed.create_chart_sheet([paths_data, referrs_data])

    exel_builed.save_exel(report_name)


class Data():
    def __init__(self):
        self._sources_info = {}
        self._date_info = {}

    def get_sources_info(self):
        return self._sources_info
    
    def get_date_info(self):
        return self._date_info

    def collect_data(self, text, key_source=URL_PATH_KEY, min_max_key=VIEWS_UNIQUE_KEY):
        reader = csv.DictReader(text.split('\n'))
        date = content.name.split('_')[0]

        self._date_info.setdefault(date, {'all_total': 0, 'all_unique': 0, 'max_source': [], 'max_val': 0, 'min_source': [], 'min_val': None})
        for row in reader:
            views_unique = int(row[VIEWS_UNIQUE_KEY])
            views_total = int(row[VIEWS_TOTAL_KEY])
            source_path = row[key_source]

            if key_source == URL_PATH_KEY:
                # let analyze only notebook or dirs with notebooks
                if Path(source_path).parent.name != 'notebooks' and '.ipynb' != Path(source_path).suffix:
                    continue
                parent_name = Path(source_path).parent.name
                if '.ipynb' == Path(source_path).suffix and self._sources_info.get(parent_name):
                    for key, item in self._sources_info[parent_name].items():
                        if key == 'all_total' or key == 'all_unique' or key == 'full_path':
                            continue
                        self._date_info[key]['all_total'] -= int(item[VIEWS_TOTAL_KEY])
                        self._date_info[key]['all_unique'] -= int(item[VIEWS_UNIQUE_KEY])
                    del self._sources_info[parent_name]
                elif '.ipynb' != Path(source_path).suffix:
                    the_same_dir = False
                    for _, info in self._sources_info.items():
                        # \openvinotoolkit\openvino_notebooks\[blob/tree]\main\notebooks\254-llm-chatbot\254-llm-chatbot.ipynb
                        if ('.ipynb' == Path(info['full_path']).suffix and
                            Path(info['full_path']).parts[0:3] == Path(source_path).parts[0:3] and
                            Path(info['full_path']).parts[4:7] == Path(source_path).parts[4:7]):
                            the_same_dir = True
                            break
                    if the_same_dir:
                        continue

                source_path = Path(row[key_source]).name


            self._sources_info.setdefault(source_path, {'all_total': 0, 'all_unique': 0, 'full_path': ''})
            self._sources_info[source_path][date] = { VIEWS_TOTAL_KEY: int(views_total), VIEWS_UNIQUE_KEY: views_unique }
            self._sources_info[source_path]['all_total'] += int(views_total)
            self._sources_info[source_path]['all_unique'] += int(views_unique)
            self._sources_info[source_path]['full_path'] = row[key_source]

            self._date_info[date]['all_total'] += int(views_total)
            self._date_info[date]['all_unique'] += int(views_unique)
            self._date_info[date]['source_name'] = content.name

            if int(row[min_max_key]) > self._date_info[date]['max_val']:
                self._date_info[date]['max_source'] = [source_path]
                self._date_info[date]['max_val'] = int(row[min_max_key])
            elif int(row[min_max_key]) == self._date_info[date]['max_val']:
                self._date_info[date]['max_source'].append(source_path)

            if self._date_info[date]['min_val'] == None:
                self._date_info[date]['min_source'] = [source_path]
                self._date_info[date]['min_val'] = int(row[min_max_key])
            elif int(row[min_max_key]) < self._date_info[date]['min_val']:
                self._date_info[date]['min_source'] = [source_path]
                self._date_info[date]['min_val'] = int(row[min_max_key])
            elif int(row[min_max_key]) == self._date_info[date]['min_val']:
                self._date_info[date]['min_source'].append(source_path)


class ExelBuiler():
    def __init__(self):
        self.workbook = openpyxl.workbook.Workbook()

        self.fail_font = openpyxl.styles.Font(color="9C0006")
        self.fail_fill = openpyxl.styles.PatternFill(start_color="FF8B8B", end_color="FF8B8B", fill_type = "solid")

        self.pass_font = openpyxl.styles.Font(color="006100")
        self.pass_fill = openpyxl.styles.PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type = "solid")

        self.title_font = openpyxl.styles.Font(bold=True)
        self.title_fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type = "solid")

    def create_content_sheet(self, sheet_name: str, data: Data):
        worksheet = self.workbook.create_sheet(sheet_name)

        sources_info = data.get_sources_info()
        date_info = data.get_date_info()
        sorted_date_info = sorted(date_info.items())

        worksheet.cell(1, 1, 'Row Labels')
        worksheet.cell(1, 1).fill = self.title_fill
        worksheet.cell(1, 1).font = self.title_font

        grand_total_date_row = len(sources_info.keys()) + 2
        worksheet.cell(grand_total_date_row, 1, 'Grand Total')
        worksheet.cell(grand_total_date_row, 1).fill = self.title_fill
        worksheet.cell(grand_total_date_row, 1).font = self.title_font

        grand_total_source_cell = len(sorted_date_info) + 2
        worksheet.cell(1, grand_total_source_cell, 'Grand Total')
        worksheet.cell(1, grand_total_source_cell).fill = self.title_fill
        worksheet.cell(1, grand_total_source_cell).font = self.title_font

        for i, date in enumerate(sorted_date_info):
            cell = i + 2
            worksheet.cell(1, cell, date[0])
            worksheet.cell(1, cell).fill = self.title_fill
            worksheet.cell(1, cell).font = self.title_font

            worksheet.cell(grand_total_date_row, cell, date[1]['all_unique'])
            worksheet.cell(grand_total_date_row, cell).fill = self.title_fill
            worksheet.cell(grand_total_date_row, cell).font = self.title_font

        row = 2
        for name, info in sources_info.items():
            worksheet.cell(row, 1, name)
            for i, date in enumerate(sorted_date_info):
                if not date[0] in info:
                    continue
                
                cell = i + 2
                if name in date[1]['max_source']:
                    worksheet.cell(row, cell).fill = self.pass_fill
                    worksheet.cell(row, cell).font = self.pass_font
                elif name in date[1]['min_source']:
                    worksheet.cell(row,cell).fill = self.fail_fill
                    worksheet.cell(row, cell).font = self.fail_font

                worksheet.cell(row, cell, info[date[0]][VIEWS_UNIQUE_KEY])
            
            worksheet.cell(row, grand_total_source_cell, info['all_unique'])

            row += 1

    def create_paths_snapshoot_sheet(self, sheet_name: str, data: Data):
        worksheet = self.workbook.create_sheet(sheet_name)

        sources_info = data.get_sources_info()
        date_info = data.get_date_info()
        sorted_date_info = sorted(date_info.items())

        for i, title in enumerate(['Source.Name', 'Url.Path', 'Total view', 'Unique view']):
            worksheet.cell(1, i + 1, title)
            worksheet.cell(1, i + 1).fill = openpyxl.styles.PatternFill(start_color="70AD47", end_color="70AD47", fill_type = "solid")
            worksheet.cell(1, i + 1).font = self.title_font

        for data_info in sorted_date_info:
            first_row = True
            for path_name, path_info in sources_info.items():
                if data_info[0] not in path_info:
                    continue
                worksheet.append((data_info[1]['source_name'], path_name, path_info[data_info[0]][VIEWS_TOTAL_KEY], path_info[data_info[0]][VIEWS_UNIQUE_KEY]))

                if first_row:
                    first_row = False
                    for cell in range(worksheet.max_column):
                        worksheet.cell(worksheet.max_row, cell + 1).fill = openpyxl.styles.PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type = "solid")


    def create_chart_sheet(self, data_list):
        worksheet = self.workbook.create_sheet('charts')
        row_offset = 0

        for data in data_list:
            sources_info = data.get_sources_info()
            date_info = data.get_date_info()
            sorted_date_info = sorted(date_info.items())

            row = row_offset + 1
            worksheet.cell(row, 1, 'Row Labels')
            worksheet.cell(row, 1).fill = self.title_fill
            worksheet.cell(row, 1).font = self.title_font
            
            for i, date in enumerate(sorted_date_info):
                worksheet.cell(row, i + 2, date[0])
                worksheet.cell(row, i + 2).fill = self.title_fill
                worksheet.cell(row, i + 2).font = self.title_font

            row = row_offset + 2
           
            for path_name, path_info in sources_info.items():
                row_data = [None] * (len(sorted_date_info) + 1)
                for i, date in enumerate(sorted_date_info):
                    if not date[0] in path_info:
                        continue
                    row_data[i+1] = path_info[date[0]][VIEWS_UNIQUE_KEY]

                if len(set(row_data)) > 1:
                    row_data[0] = path_name
                    worksheet.append(row_data)

                row += 1

            c2 = openpyxl.chart.LineChart(marker='auto')
            c2.y_axis.title = "Unique views"
            c2.x_axis.title = "Date"

            c2.height = 17
            c2.width = 35

            dates = openpyxl.chart.Reference(worksheet, min_col=1, min_row=row_offset + 2, max_row=worksheet.max_row, max_col=(len(sorted_date_info) + 1))
            c2.add_data(dates, from_rows=True, titles_from_data=True)
            dates2 = openpyxl.chart.Reference(worksheet, min_col=2, min_row=row_offset + 1, max_row=row_offset + 1, max_col=(len(sorted_date_info) + 1))
            c2.set_categories(dates2)

            for i in range(0, len(c2.series)):
                s1 = c2.series[i]
                s1.marker.symbol = "circle"
                s1.graphicalProperties.solidFill = s1.marker.graphicalProperties.line.solidFill
                s1.graphicalProperties.line.width = 1000

            worksheet.add_chart(c2, f"D{worksheet.max_row + 5 }")

            row_offset += worksheet.max_row + 50

    def save_exel(self, file_name: Path):
        for worksheet in self.workbook.worksheets:
            if worksheet.max_row <= 1:
                self.workbook.remove(worksheet)

        if len(self.workbook.worksheets) > 0:
            self.workbook.save(file_name)
            logger.info(f'Exel file with results is created: {file_name}')
        else:
            logger.info(f'Excel file hasn`t created because there is no content')

BRANCH = "github-repo-stats"
REPO = 'openvino_notebooks'
STATISTRIC_FOLDER_ROOT = 'openvinotoolkit/openvino_notebooks'

def get_contents_from_gh(token):
    logger.info(f'Get contents')
    g = Github(login_or_token=token)
    repo = g.get_repo(f"sbalandi/{REPO}")
    contents = repo.get_contents(f"{STATISTRIC_FOLDER_ROOT}/ghrs-data/snapshots", ref=BRANCH)
    return contents

def push_table_to_gh(file_name, token):
    logger.info(f'Commit final xlsx')
    g = Github(login_or_token=token)
    user = g.get_user()

    base64content = base64.b64encode(open(file_name,"rb").read()).decode("utf-8")

    current_dateTime = datetime.now().strftime("data_%m_%d_%Y_time_%H_%M_%S")

    data = {
        "message": f"add new statistics file {current_dateTime}",
        "committer": {
            "name": user.login,
            "email": user.email
        },
        "content": base64content,
        "branch": BRANCH
    }

    logger.info(f'xlsx path: {STATISTRIC_FOLDER_ROOT}/statistics/statistics_{current_dateTime}.xlsx')

    url = "{}/{}/{}/contents/{}".format(
        "https://api.github.com/repos",
        user.login,
        REPO,
        f"{STATISTRIC_FOLDER_ROOT}/statistics/statistics_{current_dateTime}.xlsx"
    )

    headers = {
                'Authorization': 'Bearer {}'.format(token),
                'X-GitHub-Api-Version': '2022-11-28',
                'Accept': 'application/vnd.github+json'
            }

    response = requests.put(
                    url,
                    headers=headers,
                    json=data
                )
    logger.info(f'Commit status {response}')

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--ghtoken")

    args = parser.parse_args()

    # contents = get_contents_files()
    contents = get_contents_from_gh(args.ghtoken)

    paths_data = Data()
    referrs_data = Data()

    for content in contents:
        logger.info(f"Analyze file {content.name}")
        text = content.decoded_content.decode('utf-8')
        if 'top_paths_snapshot' in content.name:
            paths_data.collect_data(text, key_source=URL_PATH_KEY, min_max_key=VIEWS_UNIQUE_KEY)
        elif 'top_referrers_snapshot' in content.name:
            referrs_data.collect_data(text, key_source=REFERRER_KEY, min_max_key=VIEWS_UNIQUE_KEY)

    exel_file_path = Path('report_name.xlsx')
    create_exel(paths_data, referrs_data, exel_file_path)

    push_table_to_gh(exel_file_path.as_posix(), args.ghtoken)